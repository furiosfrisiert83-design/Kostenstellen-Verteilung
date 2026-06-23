"""
Erstellt eine Excel-Vorlage für die Kostenstellen-Verteilung
mit fest integrierter Kostenstellenliste und Formeln.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# ============================================================================
# 1. NEUE ARBEITSMAPPE ERSTELLEN
# ============================================================================
wb = openpyxl.Workbook()

# ============================================================================
# 2. SHEET 1: KOSTENSTELLENLISTE (fest integriert)
# ============================================================================
ws_ks = wb.active
ws_ks.title = "Kostenstellen"

# Kopfzeile
ws_ks["A1"] = "Kostenstelle"
ws_ks["B1"] = "Bezeichnung"
ws_ks["A1"].font = Font(bold=True, color="FFFFFF")
ws_ks["B1"].font = Font(bold=True, color="FFFFFF")
ws_ks["A1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
ws_ks["B1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

# Beispiel-Kostenstellen (anonymisiert für GitHub)
beispiel_kostenstellen = [
    (1, "Kostenstelle 1"),
    (2, "Kostenstelle 2"),
    (3, "Kostenstelle 3"),
    (4, "Kostenstelle 4"),
    (5, "Kostenstelle 5"),
]

for i, (ks, bez) in enumerate(beispiel_kostenstellen, start=2):
    ws_ks[f"A{i}"] = ks
    ws_ks[f"B{i}"] = bez

# Spaltenbreite anpassen
ws_ks.column_dimensions["A"].width = 15
ws_ks.column_dimensions["B"].width = 30

# ============================================================================
# 3. SHEET 2: RECHNUNGSEINGABE
# ============================================================================
ws_eingabe = wb.create_sheet("Rechnung")

# Kopfzeile
ws_eingabe["A1"] = "Positionsbezeichnung"
ws_eingabe["B1"] = "Betrag (€)"
ws_eingabe["C1"] = "Verteilungsart"
ws_eingabe["D1"] = "Ziel-Kostenstellen"
ws_eingabe["E1"] = "Anzahl KS"

for col in ["A", "B", "C", "D", "E"]:
    ws_eingabe[f"{col}1"].font = Font(bold=True, color="FFFFFF")
    ws_eingabe[f"{col}1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

# Dropdown-Listen erstellen
from openpyxl.worksheet.datavalidation import DataValidation

# Einstellungsbereich oben im Sheet (Zeile 1-2, ausgeblendet oder als Info)
ws_eingabe["G1"] = "Einstellungen:"
ws_eingabe["G1"].font = Font(bold=True, color="FFFFFF")
ws_eingabe["G1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
ws_eingabe["H1"] = "Wert"
ws_eingabe["H1"].font = Font(bold=True, color="FFFFFF")
ws_eingabe["H1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

ws_eingabe["G2"] = "Eingabewerte sind:"
ws_eingabe["H2"] = "Netto (ohne MwSt.)"
ws_eingabe["G3"] = "Anzeige als:"
ws_eingabe["H3"] = "Netto"

# Dropdown für Eingabemodus
dv_eingabe = DataValidation(type="list", formula1='"Netto (ohne MwSt.),Brutto (mit 19% MwSt.)"', allow_blank=False)
ws_eingabe.add_data_validation(dv_eingabe)
dv_eingabe.add("H2")

# Dropdown für Anzeigemodus
dv_anzeige = DataValidation(type="list", formula1='"Netto,Brutto"', allow_blank=False)
ws_eingabe.add_data_validation(dv_anzeige)
dv_anzeige.add("H3")

dv_verteilung = DataValidation(type="list", formula1='"Einzelne Kostenstelle,Gleichmäßig auf alle,Gleichmäßig auf verwendete"', allow_blank=False)
dv_verteilung.error = "Bitte eine Verteilungsart auswählen"
dv_verteilung.errorTitle = "Ungültige Eingabe"
ws_eingabe.add_data_validation(dv_verteilung)
dv_verteilung.add("C2:C100")

# Dropdown-Liste für Ziel-Kostenstellen (aus Sheet "Kostenstellen")
# Erstelle eine kommagetrennte Liste der Kostenstellen für das Dropdown
ks_liste = [str(ks) for ks, _ in beispiel_kostenstellen]
ks_dropdown_formel = '"' + ",".join(ks_liste) + '"'

dv_kostenstellen = DataValidation(type="list", formula1=ks_dropdown_formel, allow_blank=False)
dv_kostenstellen.error = "Bitte eine gültige Kostenstelle auswählen"
dv_kostenstellen.errorTitle = "Ungültige Kostenstelle"
ws_eingabe.add_data_validation(dv_kostenstellen)
dv_kostenstellen.add("D2:D100")

# Spaltenbreite
ws_eingabe.column_dimensions["A"].width = 25
ws_eingabe.column_dimensions["B"].width = 12
ws_eingabe.column_dimensions["C"].width = 25
ws_eingabe.column_dimensions["D"].width = 40
ws_eingabe.column_dimensions["E"].width = 12

# ============================================================================
# 4. SHEET 3: ERGEBNIS
# ============================================================================
ws_ergebnis = wb.create_sheet("Ergebnis")

ws_ergebnis["A1"] = "Kostenstelle"
ws_ergebnis["B1"] = "Bezeichnung"
ws_ergebnis["C1"] = "Netto (€)"
ws_ergebnis["D1"] = "Brutto (€)"
ws_ergebnis["E1"] = "Zugewiesene Positionen"

for col in ["A", "B", "C", "D", "E"]:
    ws_ergebnis[f"{col}1"].font = Font(bold=True, color="FFFFFF")
    ws_ergebnis[f"{col}1"].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

ws_ergebnis.column_dimensions["A"].width = 15
ws_ergebnis.column_dimensions["B"].width = 30
ws_ergebnis.column_dimensions["C"].width = 15
ws_ergebnis.column_dimensions["D"].width = 15
ws_ergebnis.column_dimensions["E"].width = 40

# ============================================================================
# 5. ANLEITUNG ALS TEXT
# ===========================================================================>
anleitung = """
ANLEITUNG FÜR DIE EXCEL-VORLAGE
==================================

1. KOSTENSTELLENLISTE (Tabellenblatt "Kostenstellen")
   - Ersetze die Beispiel-Kostenstellen durch deine echte Liste
   - Spalte A: Kostenstellennummer (Zahl)
   - Spalte B: Bezeichnung (Text)

2. RECHNUNG EINGEBEN (Tabellenblatt "Rechnung")
   - Spalte A: Positionsbezeichnung (z.B. "Portokosten")
   - Spalte B: Betrag in € (z.B. 60)
   - Spalte C: Verteilungsart (Dropdown: "Einzelne Kostenstelle", "Gleichmäßig auf alle", "Gleichmäßig auf verwendete")
   - Spalte D: Ziel-Kostenstellen (z.B. "900000, 910000, 920000")
   - Spalte E: Anzahl der Kostenstellen (wird automatisch berechnet)

3. VERWENDETE KOSTENSTELLEN TRACKEN
   - Wenn du "Gleichmäßig auf verwendete" wählst, musst du die Kostenstellen
     in Spalte D eintragen. Sie werden dann für spätere Positionen gemerkt.

4. ERGEBNIS (Tabellenblatt "Ergebnis")
   - Zeigt die Verteilung pro Kostenstelle
   - Netto = Eingabewert
   - Brutto = Netto × 1,19

5. WICHTIGE HINWEISE:
   - Die Verteilungslogik muss in Excel mit Formeln oder VBA programmiert werden
   - Diese Vorlage enthält nur die Struktur, keine automatische Berechnung
   - Für eine automatische Berechnung brauchst du VBA-Makros

6. ALTERNATIVE:
   - Nutze die Streamlit-App (app.py) – sie ist bereits fertig und funktioniert
   - Die Excel-Vorlage ist nur, falls du es wirklich in Excel haben möchtest

"""

# Anleitung in neues Sheet schreiben
ws_anleitung = wb.create_sheet("Anleitung")
ws_anleitung["A1"] = anleitung
ws_anleitung.column_dimensions["A"].width = 100
ws_anleitung.row_dimensions[1].height = 400

# ============================================================================
# 6. DATEI SPEICHERN
# ============================================================================
dateiname = "kostenstellen_verteilung_excel_vorlage_anonymisiert.xlsx"
wb.save(dateiname)
print(f"✅ Excel-Vorlage erstellt: {dateiname}")
print(f"   - Sheet 1: Kostenstellen (feste Liste)")
print(f"   - Sheet 2: Rechnung (Eingabeformular)")
print(f"   - Sheet 3: Ergebnis (Ausgabe)")
print(f"   - Sheet 4: Anleitung")
print(f"\n💡 Tipp: Öffne die Datei und ersetze die Beispiel-Kostenstellen durch deine echte Liste.")